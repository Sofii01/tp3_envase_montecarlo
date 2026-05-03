from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .modelos import FilaEstado, ResultadoSimulacion


RELLENO_ENCABEZADO = PatternFill("solid", fgColor="1F4E78")
FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)


def _filas_a_dataframe(filas: list[FilaEstado]) -> pd.DataFrame:
    return pd.DataFrame([fila.a_diccionario_visible() for fila in filas])


def _auto_formatear_hoja(hoja) -> None:
    hoja.freeze_panes = "A2"

    for celda in hoja[1]:
        celda.fill = RELLENO_ENCABEZADO
        celda.font = FUENTE_ENCABEZADO
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for indice_columna, celdas_columna in enumerate(hoja.columns, start=1):
        encabezado = str(celdas_columna[0].value or "")
        ancho_maximo = min(max(len(encabezado), 12), 35)

        for celda in celdas_columna[1:60]:
            valor = "" if celda.value is None else str(celda.value)
            ancho_maximo = min(max(ancho_maximo, len(valor)), 35)
            celda.alignment = Alignment(vertical="center", wrap_text=False)

        hoja.column_dimensions[get_column_letter(indice_columna)].width = ancho_maximo + 2


def construir_excel_en_memoria(resultado: ResultadoSimulacion) -> BytesIO:
    """Genera un archivo Excel en memoria con el vector visible, fila N y resultados."""
    salida = BytesIO()

    df_visible = _filas_a_dataframe(resultado.filas_seleccionadas)
    df_final = _filas_a_dataframe([resultado.fila_final])
    df_resultados = pd.DataFrame(
        [{"Resultado": clave, "Valor": valor} for clave, valor in resultado.resultados_solicitados().items()]
    )

    with pd.ExcelWriter(salida, engine="openpyxl") as escritor:
        df_visible.to_excel(escritor, index=False, sheet_name="Vector visible")
        df_final.to_excel(escritor, index=False, sheet_name="Fila N")
        df_resultados.to_excel(escritor, index=False, sheet_name="Resultados")

        for nombre_hoja in escritor.book.sheetnames:
            _auto_formatear_hoja(escritor.book[nombre_hoja])

        hoja_resultados = escritor.book["Resultados"]
        hoja_resultados.freeze_panes = "A2"
        hoja_resultados.column_dimensions["A"].width = 62
        hoja_resultados.column_dimensions["B"].width = 22

    salida.seek(0)
    return salida
